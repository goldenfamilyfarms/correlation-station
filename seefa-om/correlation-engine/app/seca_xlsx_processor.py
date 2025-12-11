"""
SECA XLSX Processor - Parse circuit error spreadsheets and scrape detailed error reports
"""
import re
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Union
from dataclasses import dataclass, asdict
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import structlog

logger = structlog.get_logger()


@dataclass
class AffectedFile:
    """
    Detailed information about an affected file from scraping
    (Blueprint Section 4.4 - structured output dictionary)
    """
    source: str  # "orch_trace" or "other_log"
    log_file: Optional[str]  # filename or path
    traceback: Optional[str]  # full extracted traceback text
    artifact_url: Optional[str]  # URL used by Selenium
    selenium_status: str  # "ok", "artifact_not_found", "traceback_not_found", "error"


@dataclass
class CircuitError:
    """
    Circuit error data from XLSX with scraping results
    (Blueprint Section 4.4 - structured output dictionary format)
    """
    circuit_id: str
    date: str
    service_request_type: str
    product_name: str
    error_message: str
    cdnc_summary: str  # Renamed from initial_cdnc_summary
    concat_key: str  # circuit_id + "_" + date

    # Additional fields populated during scraping (Blueprint Section 4.4)
    affected_files: List[AffectedFile] = None  # List of AffectedFile objects

    # Legacy fields (for backward compatibility)
    traceback: Optional[str] = None
    log_file_path: Optional[str] = None
    categorized_error: Optional[str] = None

    def __post_init__(self):
        if self.affected_files is None:
            self.affected_files = []


class SECAXLSXProcessor:
    """Process SECA error analysis XLSX files"""

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.df: Optional[pd.DataFrame] = None
        self.errors: Dict[str, CircuitError] = {}

    def load_xlsx(self) -> pd.DataFrame:
        """Load XLSX file and extract relevant columns"""
        logger.info("Loading XLSX file", file_path=self.file_path)

        # Read XLSX with pandas
        self.df = pd.read_excel(self.file_path, engine='openpyxl')

        logger.info(
            "XLSX loaded",
            rows=len(self.df),
            columns=list(self.df.columns)
        )

        return self.df

    def parse_circuit_data(self) -> Dict[str, CircuitError]:
        """
        Extract circuit data from XLSX

        Column mapping (Blueprint Section 4.1):
        - A: Status (must be "FAIL" to process)
        - D: Circuit ID
        - E: Error Message
        - G: Product Name
        - J: Service Request Type
        - S: Date
        - W: CDNC Summary
        """
        if self.df is None:
            raise ValueError("Must call load_xlsx() first")

        # Column indices (0-based)
        STATUS_COL = 0  # A (BLUEPRINT: filter by FAIL)
        CIRCUIT_ID_COL = 3  # D
        ERROR_MSG_COL = 4   # E
        PRODUCT_NAME_COL = 6  # G
        SERVICE_TYPE_COL = 9  # J
        DATE_COL = 18  # S
        CDNC_SUMMARY_COL = 22  # W

        errors = {}
        skipped_non_fail = 0

        for idx, row in self.df.iterrows():
            try:
                # BLUEPRINT REQUIREMENT (4.1): Only process rows where Column A = FAIL
                status = str(row.iloc[STATUS_COL]).strip().upper()
                if status != "FAIL":
                    skipped_non_fail += 1
                    continue

                # Extract values
                circuit_id = str(row.iloc[CIRCUIT_ID_COL]).strip()
                date_val = row.iloc[DATE_COL]
                service_type = str(row.iloc[SERVICE_TYPE_COL]).strip()
                product_name = str(row.iloc[PRODUCT_NAME_COL]).strip()
                error_msg = str(row.iloc[ERROR_MSG_COL]).strip()
                cdnc_summary = str(row.iloc[CDNC_SUMMARY_COL]).strip()

                # Skip empty rows
                if pd.isna(circuit_id) or circuit_id == "nan":
                    continue

                # Format date
                if isinstance(date_val, pd.Timestamp):
                    date_str = date_val.strftime("%Y-%m-%d_%H-%M-%S")
                else:
                    date_str = str(date_val).strip()

                # Create concatenated key
                concat_key = f"{circuit_id}_{date_str}"

                # Create error object (Blueprint Section 4.4 format)
                error = CircuitError(
                    circuit_id=circuit_id,
                    date=date_str,
                    service_request_type=service_type,
                    product_name=product_name,
                    error_message=error_msg,
                    cdnc_summary=cdnc_summary,
                    concat_key=concat_key
                )

                errors[concat_key] = error

                logger.debug(
                    "Parsed circuit error",
                    circuit_id=circuit_id,
                    concat_key=concat_key,
                    row=idx
                )

            except Exception as e:
                logger.warning(
                    "Failed to parse row",
                    row=idx,
                    error=str(e)
                )
                continue

        self.errors = errors
        logger.info(
            "Parsed circuit errors",
            count=len(errors),
            skipped_non_fail=skipped_non_fail,
            total_rows=len(self.df)
        )

        return errors

    @staticmethod
    def extract_traceback(log_text: str) -> Optional[str]:
        """
        Extract Python traceback from log text

        Pattern:
        Traceback (most recent call last):
          File "/path/to/file.py", line 434, in run
            response = self.process()
          File "/path/to/file.py", line 58, in process
            raise Exception(f"Unable to connect to device at {device_ip}")
        Exception: Unable to connect to device at DEVICE.FQDN.COM
        """
        # Regex to find traceback block
        traceback_pattern = r'(Traceback \(most recent call last\):.*?(?:Exception|Error): .+?)(?:\n\n|\n\d{4}-\d{2}-\d{2}|$)'

        match = re.search(traceback_pattern, log_text, re.DOTALL)

        if match:
            return match.group(1).strip()

        return None

    @staticmethod
    def extract_affected_files(traceback: str) -> List[str]:
        """
        Extract file paths from traceback

        Example:
        File "/bp2/data/contexts/201/model-definitions/scripts/common_plan.py", line 434, in run
        → /bp2/data/contexts/201/model-definitions/scripts/common_plan.py
        """
        file_pattern = r'File "([^"]+)"'
        files = re.findall(file_pattern, traceback)

        # Return unique files
        return list(set(files))

    @staticmethod
    def extract_log_file_from_orch_trace(orch_trace_json: Union[Dict, List[Dict]]) -> Optional[str]:
        """
        Extract log file name from orchestration trace JSON

        Example:
        {
            "timestamp": "01:42:08",
            "state": "FAILED",
            "log_file": "plan-script-4cbd1db2-...-2025-12-10T01:41:08.538Z",
            ...
        }
        """
        if isinstance(orch_trace_json, dict):
            # Look for first FAILED entry
            if orch_trace_json.get("state") == "FAILED":
                return orch_trace_json.get("log_file")

        # Handle list of trace entries
        if isinstance(orch_trace_json, list):
            for entry in orch_trace_json:
                if entry.get("state") == "FAILED":
                    return entry.get("log_file")

        return None

    def generate_reformatted_xlsx(self, output_path: str) -> str:
        """
        Generate reformatted XLSX with:
        - Filtered by FAIL in column A
        - Hidden columns: B, C, H, K, M, N, D, P, Q, R, S, T, X, AE, AF
        - Column E sorted A-Z
        - Identical values in E highlighted with same color
        """
        if self.df is None:
            raise ValueError("Must call load_xlsx() first")

        logger.info("Generating reformatted XLSX", output_path=output_path)

        # Load workbook
        wb = load_workbook(self.file_path)
        ws = wb.active

        # Filter by FAIL in column A
        # (This is display filtering - actual filtering done in Excel UI)

        # Sort by column E (error message) - alphabetically
        # Convert to list for sorting
        data_rows = list(ws.iter_rows(min_row=2))  # Skip header

        # Sort by column E (index 4)
        sorted_rows = sorted(
            data_rows,
            key=lambda row: str(row[4].value or "").strip().lower()
        )

        # Clear existing data (keep header)
        ws.delete_rows(2, ws.max_row)

        # Write sorted data back
        for row_idx, row in enumerate(sorted_rows, start=2):
            for col_idx, cell in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx).value = cell.value

        # Color-code identical values in column E
        error_colors = {}
        color_palette = [
            "FFB6D7A8",  # Light green
            "FFFCE5CD",  # Light orange
            "FFD9EAD3",  # Light olive
            "FFC9DAF8",  # Light blue
            "FFD0E0E3",  # Light gray-blue
            "FFF4CCCC",  # Light red
            "FFFCE8B2",  # Light yellow
            "FFE6B8AF",  # Light brown
            "FFB4A7D6",  # Light purple
            "FFD5A6BD",  # Light pink
        ]

        current_color_idx = 0

        for row_idx in range(2, ws.max_row + 1):
            error_msg = str(ws.cell(row=row_idx, column=5).value or "").strip()

            if not error_msg:
                continue

            # Assign color to this error message
            if error_msg not in error_colors:
                error_colors[error_msg] = color_palette[current_color_idx % len(color_palette)]
                current_color_idx += 1

            # Apply color to entire row
            fill_color = error_colors[error_msg]
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    start_color=fill_color,
                    end_color=fill_color,
                    fill_type="solid"
                )

        # Hide columns: B, C, H, K, M, N, D, P, Q, R, S, T, X, AE, AF
        # Column indices (1-based): B=2, C=3, D=4, H=8, K=11, M=13, N=14, P=16, Q=17, R=18, S=19, T=20, X=24, AE=31, AF=32
        hidden_cols = [2, 3, 4, 8, 11, 13, 14, 16, 17, 18, 19, 20, 24, 31, 32]

        for col_idx in hidden_cols:
            if col_idx <= ws.max_column:
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].hidden = True

        # Save workbook
        wb.save(output_path)

        logger.info("Reformatted XLSX saved", output_path=output_path)

        return output_path

    def to_dict(self) -> Dict:
        """Convert all circuit errors to dictionary"""
        return {
            key: asdict(error)
            for key, error in self.errors.items()
        }
